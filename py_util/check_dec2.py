import openpyxl
from datetime import datetime

wb1 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 31.03.25.xlsx', data_only=True)
ws = wb1['заб-ть с 15.12.2020']

# Дата в J (индекс 9), всего=K(10), за сутки=L(11), летальных всего=P(15), за сутки=Q(16)
# Проверим строку 8
print("Строка 8, колонки J-Q:")
for cell in ws[8]:
    if cell.value is not None:
        print(f"  {cell.column_letter}: {cell.value}")

# Попробуем прочитать напрямую
print("\nПрямое чтение строк 8-12, колонки J(10)-Q(17):")
for row in ws.iter_rows(min_row=8, max_row=12, min_col=10, max_col=17, values_only=True):
    print(f"  J={row[0]}, K={row[1]}, L={row[2]}, P={row[6]}, Q={row[7]}")
