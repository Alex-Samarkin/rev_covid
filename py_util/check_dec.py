import openpyxl
from datetime import datetime

wb1 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 31.03.25.xlsx', data_only=True)

# Проверим лист "заб-ть с 15.12.2020" — данные в колонке J начиная со строки 8
ws = wb1['заб-ть с 15.12.2020']
print(f"Лист 'заб-ть с 15.12.2020': {ws.max_row} строк x {ws.max_column} колонок")

# Проверим строки 1-20 в колонках A-L
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=12, values_only=False)):
    vals = [(cell.value, cell.column_letter) for cell in row if cell.value is not None]
    if vals:
        print(f"  Строка {i+1}: {vals}")

# Попробуем найти даты в колонке J
print("\nПоиск дат в колонке J (строки 1-50):")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, min_col=10, max_col=10, values_only=False)):
    cell = row[0]
    if cell.value is not None:
        print(f"  Строка {i+1}, J: {cell.value} (type={type(cell.value).__name__})")
