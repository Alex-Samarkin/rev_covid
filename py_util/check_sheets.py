import openpyxl
from datetime import datetime

wb1 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 31.03.25.xlsx', data_only=True)

# Проверим листы "летальность" — у них другая структура
for sheet_name in wb1.sheetnames:
    if 'летальн' in sheet_name.lower():
        ws = wb1[sheet_name]
        print(f"\n=== Лист: '{sheet_name}' ===")
        print(f"Размеры: {ws.max_row} строк x {ws.max_column} колонок")
        # Покажем первые 10 строк
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=False)):
            vals = [(cell.value, cell.column_letter) for cell in row if cell.value is not None]
            if vals:
                print(f"  Строка {i+1}: {vals}")

# Проверим пустые листы
for sheet_name in ['заб-ть с 15.12.2020', 'летальность с 15.12.2020']:
    if sheet_name in wb1.sheetnames:
        ws = wb1[sheet_name]
        print(f"\n=== Лист: '{sheet_name}' ===")
        print(f"Размеры: {ws.max_row} строк x {ws.max_column} колонок")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 15), values_only=False)):
            vals = [(cell.value, cell.column_letter) for cell in row if cell.value is not None]
            if vals:
                print(f"  Строка {i+1}: {vals}")
