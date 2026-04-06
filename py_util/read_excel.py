import openpyxl
import os

files = [
    r'C:\Users\alex_\Documents\rev_covid\COVID-19 31.03.25.xlsx',
    r'C:\Users\alex_\Documents\rev_covid\COVID-19 2090 07.04-22.12.2025.xlsx',
    r'C:\Users\alex_\Documents\rev_covid\ВП 2024-2025.xlsx'
]

for fpath in files:
    print(f'\n===== {os.path.basename(fpath)} =====')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n--- Лист: {sheet_name} ---')
        print(f'Размеры: {ws.max_row} строк x {ws.max_column} колонок')
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=False):
            vals = [(cell.value, cell.coordinate) for cell in row if cell.value is not None]
            if vals:
                print(vals)
