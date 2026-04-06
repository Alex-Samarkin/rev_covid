import openpyxl
import csv
from datetime import datetime, timedelta

# ============================================================
# ШАГ 1: Определяем все листы
# ============================================================
print("=" * 80)
print("ШАГ 1: Определение всех листов")
print("=" * 80)

wb1 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 31.03.25.xlsx', data_only=True)
print(f"\nФАЙЛ 1 листы: {wb1.sheetnames}")

wb2 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 2090 07.04-22.12.2025.xlsx', data_only=True)
print(f"\nФАЙЛ 2 листы: {wb2.sheetnames}")

wb3 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\ВП 2024-2025.xlsx', data_only=True)
print(f"\nФАЙЛ 3 листы: {wb3.sheetnames}")

# ============================================================
# ШАГ 2: Извлечение COVID-19
# ============================================================
print("\n" + "=" * 80)
print("ШАГ 2: Извлечение данных COVID-19")
print("=" * 80)

covid_data = {}  # {date_str: {total, daily, deaths_total, deaths_daily}}

# Файл 1: все листы, ищем данные о COVID-19 (не ВП)
for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    # Пропускаем листы с ВП
    if 'ВП' in sheet_name.upper() or 'пневмон' in sheet_name.lower():
        print(f"  Пропуск листа ВП: '{sheet_name}'")
        continue
    
    print(f"\n  Лист: '{sheet_name}'")
    count = 0
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=60, values_only=True):
        date_val = row[0]  # A - дата
        total_cases = row[1]  # B - всего
        daily_cases = row[2]  # C - за сутки
        total_deaths = row[6]  # G - летальных всего
        daily_deaths = row[7]  # H - летальных за сутки
        
        if date_val is None:
            continue
        # Проверяем, что это дата
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            try:
                dt = datetime.strptime(str(date_val)[:10], '%Y-%m-%d')
                date_str = dt.strftime('%Y-%m-%d')
            except:
                continue
        
        # Проверяем, что есть хоть какие-то данные
        if total_cases is None and daily_cases is None:
            continue
            
        covid_data[date_str] = {
            'total': total_cases if total_cases is not None else 'MD',
            'daily': daily_cases if daily_cases is not None else 'MD',
            'deaths_total': total_deaths if total_deaths is not None else 'MD',
            'deaths_daily': daily_deaths if daily_deaths is not None else 'MD',
        }
        count += 1
    print(f"    Извлечено записей: {count}")

# Файл 2: COVID-19 2025
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    print(f"\n  Лист файла 2: '{sheet_name}'")
    count = 0
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=60, values_only=True):
        date_val = row[0]
        total_cases = row[1]
        daily_cases = row[2]
        total_deaths = row[6]
        daily_deaths = row[7]
        
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            try:
                dt = datetime.strptime(str(date_val)[:10], '%Y-%m-%d')
                date_str = dt.strftime('%Y-%m-%d')
            except:
                continue
        
        if total_cases is None and daily_cases is None:
            continue
        
        covid_data[date_str] = {
            'total': total_cases if total_cases is not None else 'MD',
            'daily': daily_cases if daily_cases is not None else 'MD',
            'deaths_total': total_deaths if total_deaths is not None else 'MD',
            'deaths_daily': daily_deaths if daily_deaths is not None else 'MD',
        }
        count += 1
    print(f"    Извлечено записей: {count}")

print(f"\n  Всего уникальных дат COVID-19: {len(covid_data)}")
if covid_data:
    sorted_dates = sorted(covid_data.keys())
    print(f"  Диапазон: {sorted_dates[0]} — {sorted_dates[-1]}")

# ============================================================
# ШАГ 3: Извлечение ВП
# ============================================================
print("\n" + "=" * 80)
print("ШАГ 3: Извлечение данных ВП (внебольничные пневмонии)")
print("=" * 80)

vp_data = {}  # {date_str: {total, weekly, deaths_total, ...}}

for sheet_name in wb3.sheetnames:
    ws = wb3[sheet_name]
    print(f"\n  Лист: '{sheet_name}'")
    count = 0
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=50, values_only=True):
        date_val = row[0]  # A - дата
        vp_total = row[1]  # B - всего ВП накоп.
        vp_weekly = row[2]  # C - за неделю
        deaths_total = row[16]  # Q - летальные накоп.
        deaths_weekly = row[18] if len(row) > 18 else None  # S - летальные за неделю
        
        if date_val is None:
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            try:
                dt = datetime.strptime(str(date_val)[:10], '%Y-%m-%d')
                date_str = dt.strftime('%Y-%m-%d')
            except:
                continue
        
        if vp_total is None and vp_weekly is None:
            continue
        
        vp_data[date_str] = {
            'vp_total': vp_total if vp_total is not None else 'MD',
            'vp_weekly': vp_weekly if vp_weekly is not None else 'MD',
            'deaths_total': deaths_total if deaths_total is not None else 'MD',
            'deaths_weekly': deaths_weekly if deaths_weekly is not None else 'MD',
        }
        count += 1
    print(f"    Извлечено записей: {count}")

print(f"\n  Всего уникальных дат ВП: {len(vp_data)}")
if vp_data:
    sorted_dates_vp = sorted(vp_data.keys())
    print(f"  Диапазон: {sorted_dates_vp[0]} — {sorted_dates_vp[-1]}")

# ============================================================
# ШАГ 4: Заполнение пропусков и сохранение CSV
# ============================================================
print("\n" + "=" * 80)
print("ШАГ 4: Заполнение пропусков и сохранение CSV")
print("=" * 80)

# --- COVID-19 CSV ---
if covid_data:
    sorted_covid = sorted(covid_data.keys())
    first_date = datetime.strptime(sorted_covid[0], '%Y-%m-%d')
    last_date = datetime.strptime(sorted_covid[-1], '%Y-%m-%d')
    
    covid_csv_path = r'C:\Users\alex_\Documents\rev_covid\covid19_daily.csv'
    with open(covid_csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['Дата', 'Всего (накоп.)', 'За сутки', 'Летальных всего', 'Летальных за сутки'])
        
        current = first_date
        filled = 0
        missing = 0
        while current <= last_date:
            date_str = current.strftime('%Y-%m-%d')
            if date_str in covid_data:
                d = covid_data[date_str]
                writer.writerow([date_str, d['total'], d['daily'], d['deaths_total'], d['deaths_daily']])
                filled += 1
            else:
                writer.writerow([date_str, 'MD', 'MD', 'MD', 'MD'])
                missing += 1
            current += timedelta(days=1)
    
    total_days = (last_date - first_date).days + 1
    print(f"\n  COVID-19: {covid_csv_path}")
    print(f"  Диапазон: {sorted_covid[0]} — {sorted_covid[-1]} ({total_days} дней)")
    print(f"  Дней с данными: {filled}, пропусков (MD): {missing}")

# --- ВП CSV ---
if vp_data:
    sorted_vp = sorted(vp_data.keys())
    first_date_vp = datetime.strptime(sorted_vp[0], '%Y-%m-%d')
    last_date_vp = datetime.strptime(sorted_vp[-1], '%Y-%m-%d')
    
    vp_csv_path = r'C:\Users\alex_\Documents\rev_covid\vp_daily.csv'
    with open(vp_csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['Дата', 'ВП всего (накоп.)', 'ВП за период', 'Летальных всего', 'Летальных за период'])
        
        current = first_date_vp
        filled = 0
        missing = 0
        while current <= last_date_vp:
            date_str = current.strftime('%Y-%m-%d')
            if date_str in vp_data:
                d = vp_data[date_str]
                writer.writerow([date_str, d['vp_total'], d['vp_weekly'], d['deaths_total'], d['deaths_weekly']])
                filled += 1
            else:
                writer.writerow([date_str, 'MD', 'MD', 'MD', 'MD'])
                missing += 1
            current += timedelta(days=1)
    
    total_days_vp = (last_date_vp - first_date_vp).days + 1
    print(f"\n  ВП: {vp_csv_path}")
    print(f"  Диапазон: {sorted_vp[0]} — {sorted_vp[-1]} ({total_days_vp} дней)")
    print(f"  Дней с данными: {filled}, пропусков (MD): {missing}")

print("\n" + "=" * 80)
print("ГОТОВО!")
print("=" * 80)
