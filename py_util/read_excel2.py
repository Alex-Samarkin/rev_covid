import openpyxl
import os

# ===== ФАЙЛ 1: COVID-19 31.03.25.xlsx =====
print("=" * 80)
print("ФАЙЛ 1: COVID-19 31.03.25.xlsx")
print("Заболеваемость COVID-19 с 15.04 по 31.03.25 (накопленным итогом и за сутки)")
print("=" * 80)

wb = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 31.03.25.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Лист: {sheet_name} ---")
    # Данные начинаются с 8 строки, колонка A - дата, B - всего, C - за сутки
    print(f"{'Дата':<15} {'Всего (накоп.)':>15} {'За сутки':>12} {'Летальных всего':>18} {'Летальных за сутки':>20} {'Госпитализировано всего':>24}")
    print("-" * 110)
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=55, values_only=True):
        date_val = row[0]  # A - дата
        total_cases = row[1]  # B - всего заболевших
        daily_cases = row[2]  # C - за сутки
        total_deaths = row[6]  # G - летальных всего
        daily_deaths = row[7]  # H - летальных за сутки
        total_hosp = row[51]  # AZ - госпитализировано всего
        
        if date_val is not None:
            date_str = str(date_val)[:10] if hasattr(date_val, 'strftime') else str(date_val)
            print(f"{date_str:<15} {str(total_cases):>15} {str(daily_cases):>12} {str(total_deaths):>18} {str(daily_deaths):>20} {str(total_hosp):>24}")


# ===== ФАЙЛ 2: COVID-19 2090 07.04-22.12.2025.xlsx =====
print("\n" + "=" * 80)
print("ФАЙЛ 2: COVID-19 2090 07.04-22.12.2025.xlsx")
print("Заболеваемость COVID-19 с 07.04 по 22.12.2025")
print("=" * 80)

wb2 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 2090 07.04-22.12.2025.xlsx', data_only=True)
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    print(f"\n--- Лист: {sheet_name} ---")
    print(f"{'Дата':<15} {'Всего (накоп.)':>15} {'За сутки':>12} {'Летальных всего':>18} {'Летальных за сутки':>20}")
    print("-" * 85)
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=55, values_only=True):
        date_val = row[0]
        total_cases = row[1]
        daily_cases = row[2]
        total_deaths = row[6]
        daily_deaths = row[7]
        
        if date_val is not None:
            date_str = str(date_val)[:10] if hasattr(date_val, 'strftime') else str(date_val)
            print(f"{date_str:<15} {str(total_cases):>15} {str(daily_cases):>12} {str(total_deaths):>18} {str(daily_deaths):>20}")


# ===== ФАЙЛ 3: ВП 2024-2025.xlsx =====
print("\n" + "=" * 80)
print("ФАЙЛ 3: ВП 2024-2025.xlsx")
print("Внебольничные пневмонии (ВП) — еженедельные данные")
print("=" * 80)

wb3 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\ВП 2024-2025.xlsx', data_only=True)
for sheet_name in wb3.sheetnames:
    ws = wb3[sheet_name]
    print(f"\n--- Лист: {sheet_name} ---")
    print(f"{'Дата (неделя)':<18} {'ВП всего (накоп.)':>20} {'ВП за неделю':>15} {'Летальные (накоп.)':>20} {'COVID-19 (накоп.)':>20} {'Грипп A/B (накоп.)':>20}")
    print("-" * 115)
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=50, values_only=True):
        date_val = row[0]
        vp_total = row[1]  # B - всего ВП накопительным
        vp_weekly = row[2]  # C - за неделю
        deaths_total = row[16]  # Q - летальные накоп.
        covid = row[23]  # X - COVID-19 накоп.
        flu = row[25]  # Z - Грипп A/B накоп.
        
        if date_val is not None:
            date_str = str(date_val)[:10] if hasattr(date_val, 'strftime') else str(date_val)
            print(f"{date_str:<18} {str(vp_total):>20} {str(vp_weekly):>15} {str(deaths_total):>20} {str(covid):>20} {str(flu):>20}")
