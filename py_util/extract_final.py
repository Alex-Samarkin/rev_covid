import openpyxl
import csv
from datetime import datetime, timedelta

# ============================================================
# Вспомогательные функции
# ============================================================
def parse_date(val):
    """Преобразует значение в дату (str YYYY-MM-DD) или None"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()[:10]
    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%Y-%m-%d')
    except ValueError:
        return None

# ============================================================
# ШАГ 1: Загрузка книг
# ============================================================
print("Загрузка файлов...")
wb1 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 31.03.25.xlsx', data_only=True)
wb2 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\COVID-19 2090 07.04-22.12.2025.xlsx', data_only=True)
wb3 = openpyxl.load_workbook(r'C:\Users\alex_\Documents\rev_covid\ВП 2024-2025.xlsx', data_only=True)

# ============================================================
# ШАГ 2: Извлечение COVID-19
# ============================================================
print("\nИзвлечение данных COVID-19...")

# covid_data[date] = {total, daily, deaths_total, deaths_daily}
covid_data = {}

def extract_covid_from_sheet(ws, date_col_letter, data_start_row, 
                              total_col_offset, daily_col_offset,
                              deaths_total_col_offset, deaths_daily_col_offset):
    """
    Извлекает данные COVID из листа с известной структурой.
    date_col_letter: буква колонки с датой
    data_start_row: номер строки начала данных
    *_col_offset: смещение колонки относительно date_col (A=0, B=1, ...)
    """
    date_col_idx = ord(date_col_letter.upper()) - ord('A')  # 0-based
    total_col = date_col_idx + total_col_offset
    daily_col = date_col_idx + daily_col_offset
    deaths_total_col = date_col_idx + deaths_total_col_offset
    deaths_daily_col = date_col_idx + deaths_daily_col_offset
    
    all_cols = [date_col_idx, total_col, daily_col, deaths_total_col, deaths_daily_col]
    min_c = min(all_cols)
    max_c = max(all_cols)
    
    count = 0
    # iter_rows использует 1-based колонки, row_cells — кортеж Cell от min_col до max_col
    # Индекс в row_cells: cell_index = absolute_col_idx - min_c
    for row_cells in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, 
                            min_col=min_c + 1,  # 1-based
                            max_col=max_c + 1,
                            values_only=False):
        # Относительные индексы в row_cells
        di = date_col_idx - min_c
        ti = total_col - min_c
        dci = daily_col - min_c
        dti = deaths_total_col - min_c
        ddi = deaths_daily_col - min_c
        
        date_val = row_cells[di].value if di < len(row_cells) else None
        ds = parse_date(date_val)
        if ds is None:
            continue
        
        total = row_cells[ti].value if ti < len(row_cells) else None
        daily = row_cells[dci].value if dci < len(row_cells) else None
        dt = row_cells[dti].value if dti < len(row_cells) else None
        dd = row_cells[ddi].value if ddi < len(row_cells) else None
        
        if total is None and daily is None and dt is None and dd is None:
            continue
        
        # Если запись уже есть — берём наиболее полные данные
        if ds in covid_data:
            existing = covid_data[ds]
            if total is not None:
                existing['total'] = total
            if daily is not None:
                existing['daily'] = daily
            if dt is not None:
                existing['deaths_total'] = dt
            if dd is not None:
                existing['deaths_daily'] = dd
        else:
            covid_data[ds] = {
                'total': total if total is not None else 'MD',
                'daily': daily if daily is not None else 'MD',
                'deaths_total': dt if dt is not None else 'MD',
                'deaths_daily': dd if dd is not None else 'MD',
            }
        count += 1
    return count

# --- Файл 1: ранние листы (апрель-ноябрь 2020) ---
# Структура: дата в A, всего=B(1), за сутки=C(2), летальных всего=G(6), за сутки=H(7)
early_sheets = [
    'Забо-сть с 15.04. по 17.05.20',
    'Забо-сть с 17.05. по 18.06.2020',
    'Забол-сть с 19.06.по 19.07.2020',
    'Забол-сть с 20.07.по25.08.2020 ',
    'Забол-сть с 25.08.по21.09.2020',
    'заб-ть с 22.09. по 20.10.2020',
    'заб-ть с 21.10 по 23.11.2020',
    'заб-ть с 23.11.20 по 14.12.2020',
]
for sn in early_sheets:
    if sn in wb1.sheetnames:
        ws = wb1[sn]
        c = extract_covid_from_sheet(ws, 'A', 8, 1, 2, 6, 7)
        print(f"  '{sn}': {c} записей")

# --- Файл 1: листы летальности (апрель-ноябрь 2020) ---
# Структура: дата в A, летальных всего=B(1), за сутки=C(2)
lethality_early = [
    'летальность с 17.05.по 18.06.20',
    'летальность с18.06.по19.07.2020',
    'летальность с 20.07.по25.08.202',
    'летальность с25.08.по21.09.2020',
    'летальность с 22.09.по 20.10.20',
    'летальность с 21.10 по 23.11.20',
    'летальность с 23.11по14.12.2020',
]
for sn in lethality_early:
    if sn in wb1.sheetnames:
        ws = wb1[sn]
        # Данные начинаются со строки 11 (после заголовков)
        date_col_idx = 0  # A
        deaths_total_col = 1  # B
        deaths_daily_col = 2  # C
        count = 0
        for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=0, max_col=3, values_only=True):
            ds = parse_date(row[0])
            if ds is None:
                continue
            dt = row[1] if len(row) > 1 else None
            dd = row[2] if len(row) > 2 else None
            if dt is None and dd is None:
                continue
            if ds in covid_data:
                if dt is not None:
                    covid_data[ds]['deaths_total'] = dt
                if dd is not None:
                    covid_data[ds]['deaths_daily'] = dd
            else:
                covid_data[ds] = {
                    'total': 'MD',
                    'daily': 'MD',
                    'deaths_total': dt if dt is not None else 'MD',
                    'deaths_daily': dd if dd is not None else 'MD',
                }
            count += 1
        print(f"  '{sn}': {count} записей (летальность)")

# --- Файл 1: листы с декабря 2020 ---
# Заболеваемость: дата в J(9), всего=K(+1), за сутки=L(+2), летальных всего=P(+6), за сутки=Q(+7)
if 'заб-ть с 15.12.2020' in wb1.sheetnames:
    ws = wb1['заб-ть с 15.12.2020']
    c = extract_covid_from_sheet(ws, 'J', 8, 1, 2, 6, 7)
    print(f"  'заб-ть с 15.12.2020': {c} записей")

# Летальность: дата в H(7), всего=I(+1), за сутки=J(+2)
if 'летальность с 15.12.2020' in wb1.sheetnames:
    ws = wb1['летальность с 15.12.2020']
    date_col_idx = 7  # H
    deaths_total_col = 8  # I
    deaths_daily_col = 9  # J
    count = 0
    for row in ws.iter_rows(min_row=16, max_row=ws.max_row, min_col=8, max_col=10, values_only=True):
        ds = parse_date(row[0])
        if ds is None:
            continue
        dt = row[1] if len(row) > 1 else None
        dd = row[2] if len(row) > 2 else None
        if dt is None and dd is None:
            continue
        if ds in covid_data:
            if dt is not None:
                covid_data[ds]['deaths_total'] = dt
            if dd is not None:
                covid_data[ds]['deaths_daily'] = dd
        else:
            covid_data[ds] = {
                'total': 'MD',
                'daily': 'MD',
                'deaths_total': dt if dt is not None else 'MD',
                'deaths_daily': dd if dd is not None else 'MD',
            }
        count += 1
    print(f"  'летальность с 15.12.2020': {count} записей (летальность)")

# --- Файл 2: COVID-19 2025 ---
for sheet_name in wb2.sheetnames:
    ws = wb2[sheet_name]
    # Структура как у ранних листов: A=дата, B=всего, C=за сутки, G=летальных всего, H=летальных за сутки
    # Но данные могут быть в других колонках — проверим
    count = 0
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=0, max_col=10, values_only=True):
        ds = parse_date(row[0])
        if ds is None:
            continue
        total = row[1] if len(row) > 1 else None
        daily = row[2] if len(row) > 2 else None
        dt = row[6] if len(row) > 6 else None
        dd = row[7] if len(row) > 7 else None
        if total is None and daily is None:
            continue
        if ds in covid_data:
            if total is not None:
                covid_data[ds]['total'] = total
            if daily is not None:
                covid_data[ds]['daily'] = daily
            if dt is not None:
                covid_data[ds]['deaths_total'] = dt
            if dd is not None:
                covid_data[ds]['deaths_daily'] = dd
        else:
            covid_data[ds] = {
                'total': total if total is not None else 'MD',
                'daily': daily if daily is not None else 'MD',
                'deaths_total': dt if dt is not None else 'MD',
                'deaths_daily': dd if dd is not None else 'MD',
            }
        count += 1
    print(f"  Файл 2 '{sheet_name}': {count} записей")

print(f"\n  Всего уникальных дат COVID-19: {len(covid_data)}")
if covid_data:
    sorted_dates = sorted(covid_data.keys())
    print(f"  Диапазон: {sorted_dates[0]} — {sorted_dates[-1]}")

# ============================================================
# ШАГ 3: Извлечение ВП
# ============================================================
print("\nИзвлечение данных ВП...")

vp_data = {}
for sheet_name in wb3.sheetnames:
    ws = wb3[sheet_name]
    count = 0
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=0, max_col=25, values_only=True):
        ds = parse_date(row[0])
        if ds is None:
            continue
        vp_total = row[1] if len(row) > 1 else None
        vp_weekly = row[2] if len(row) > 2 else None
        deaths_total = row[16] if len(row) > 16 else None
        deaths_weekly = row[18] if len(row) > 18 else None
        
        if vp_total is None and vp_weekly is None:
            continue
        
        if ds in vp_data:
            if vp_total is not None:
                vp_data[ds]['vp_total'] = vp_total
            if vp_weekly is not None:
                vp_data[ds]['vp_weekly'] = vp_weekly
            if deaths_total is not None:
                vp_data[ds]['deaths_total'] = deaths_total
            if deaths_weekly is not None:
                vp_data[ds]['deaths_weekly'] = deaths_weekly
        else:
            vp_data[ds] = {
                'vp_total': vp_total if vp_total is not None else 'MD',
                'vp_weekly': vp_weekly if vp_weekly is not None else 'MD',
                'deaths_total': deaths_total if deaths_total is not None else 'MD',
                'deaths_weekly': deaths_weekly if deaths_weekly is not None else 'MD',
            }
        count += 1
    print(f"  '{sheet_name}': {count} записей")

print(f"\n  Всего уникальных дат ВП: {len(vp_data)}")
if vp_data:
    sorted_vp = sorted(vp_data.keys())
    print(f"  Диапазон: {sorted_vp[0]} — {sorted_vp[-1]}")

# ============================================================
# ШАГ 4: Сохранение CSV с заполнением пропусков
# ============================================================
print("\nСохранение CSV...")

# --- COVID-19 ---
if covid_data:
    sorted_covid = sorted(covid_data.keys())
    first_date = datetime.strptime(sorted_covid[0], '%Y-%m-%d')
    last_date = datetime.strptime(sorted_covid[-1], '%Y-%m-%d')
    
    covid_csv_path = r'C:\Users\alex_\Documents\rev_covid\covid19_daily.csv'
    with open(covid_csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['Дата', 'Всего (накоп.)', 'За сутки', 'Летальных всего', 'Летальных за сутки'])
        
        current = first_date
        filled = 0
        missing = 0
        while current <= last_date:
            date_str = current.strftime('%Y-%m-%d')
            if date_str in covid_data:
                d = covid_data[date_str]
                writer.writerow([date_str, d['total'], d['daily'], d['deaths_total'], d['deaths_daily']])
                filled += 1
            else:
                writer.writerow([date_str, 'MD', 'MD', 'MD', 'MD'])
                missing += 1
            current += timedelta(days=1)
    
    total_days = (last_date - first_date).days + 1
    print(f"\n  COVID-19: {covid_csv_path}")
    print(f"  Диапазон: {sorted_covid[0]} — {sorted_covid[-1]} ({total_days} дней)")
    print(f"  Дней с данными: {filled}, пропусков (MD): {missing}")

# --- ВП ---
if vp_data:
    sorted_vp = sorted(vp_data.keys())
    first_date_vp = datetime.strptime(sorted_vp[0], '%Y-%m-%d')
    last_date_vp = datetime.strptime(sorted_vp[-1], '%Y-%m-%d')
    
    vp_csv_path = r'C:\Users\alex_\Documents\rev_covid\vp_daily.csv'
    with open(vp_csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['Дата', 'ВП всего (накоп.)', 'ВП за период', 'Летальных всего', 'Летальных за период'])
        
        current = first_date_vp
        filled = 0
        missing = 0
        while current <= last_date_vp:
            date_str = current.strftime('%Y-%m-%d')
            if date_str in vp_data:
                d = vp_data[date_str]
                writer.writerow([date_str, d['vp_total'], d['vp_weekly'], d['deaths_total'], d['deaths_weekly']])
                filled += 1
            else:
                writer.writerow([date_str, 'MD', 'MD', 'MD', 'MD'])
                missing += 1
            current += timedelta(days=1)
    
    total_days_vp = (last_date_vp - first_date_vp).days + 1
    print(f"\n  ВП: {vp_csv_path}")
    print(f"  Диапазон: {sorted_vp[0]} — {sorted_vp[-1]} ({total_days_vp} дней)")
    print(f"  Дней с данными: {filled}, пропусков (MD): {missing}")

# Покажем первые и последние строки каждого CSV
print("\n--- Первые 5 строк covid19_daily.csv ---")
with open(covid_csv_path, 'r', encoding='utf-8-sig') as f:
    for i, line in enumerate(f):
        if i < 6:
            print(line.rstrip())

print("\n--- Последние 5 строк covid19_daily.csv ---")
with open(covid_csv_path, 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()
    for line in lines[-5:]:
        print(line.rstrip())

print("\n--- Первые 5 строк vp_daily.csv ---")
with open(vp_csv_path, 'r', encoding='utf-8-sig') as f:
    for i, line in enumerate(f):
        if i < 6:
            print(line.rstrip())

print("\n--- Последние 5 строк vp_daily.csv ---")
with open(vp_csv_path, 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()
    for line in lines[-5:]:
        print(line.rstrip())

print("\n" + "=" * 80)
print("ГОТОВО!")
print("=" * 80)
